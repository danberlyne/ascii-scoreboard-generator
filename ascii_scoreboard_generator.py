#!/usr/bin/env python3
# ascii_scoreboard_generator.py - Generates a set of ASCII scoreboards from a spreadsheet of voting data.

import openpyxl
from pathlib import Path

def main():
    wb = openpyxl.load_workbook('voting_data.xlsx')
    ws = wb.active
    voters = [cell for cell in ws['1'] if cell.value != '' and cell.value != None]
    competitors = [cell for cell in ws['A'] if cell.value != '' and cell.value != None]

    max_name_length = max(max(len(voter.value) for voter in voters), max(len(competitor.value) for competitor in competitors))
    scoreboard_width = max_name_length + 22
    horizontal_line = ' ' + '_' * scoreboard_width + ' '
    dead_space = '|' + ' ' * scoreboard_width + '|'
    header = '|' + '%s' + '|'
    competitor_scorecard = '| ' + '%s' + ' ' + '%s' + '%s' + ' | ' + '%s' + ' |'
    footer = '|' + '_' * scoreboard_width + '|'
    # `template` has the following string interpolation:
    # - `voter`;
    # - `position`, `competitor`, `vote`, and `running_total` for each competitor.
    template = f'{horizontal_line}\n{dead_space}\n{header}\n{footer}\n{dead_space}\n' + f'{competitor_scorecard}\n' * len(competitors) + footer

    running_totals = {competitor: 0 for competitor in competitors}
    for i, voter in enumerate(voters):
        votes = {competitor: ws[f'{voter.column_letter}{competitor.row}'].value for competitor in competitors}
        for competitor in competitors:
            if votes[competitor] == None:
                votes[competitor] = ''
        scoreboard_data = (f'NOW VOTING: {voter.value} ({i+1}/{len(voters)})'.center(scoreboard_width),)
        for competitor in competitors:
            if type(votes[competitor]) == int:
                running_totals[competitor] += votes[competitor]
        running_totals = dict(sorted(running_totals.items(), key=lambda x:x[1], reverse=True))
        for j, competitor in enumerate(running_totals):
            scoreboard_data += (f'{j + 1}'.zfill(2), competitor.value.ljust(scoreboard_width - 13), str(votes[competitor]).rjust(2), str(running_totals[competitor]).zfill(3))
        scoreboard = template % scoreboard_data
        file = open(Path('Scoreboards', f'{i + 1}'.zfill(2) + f' {voter.value}.txt'), 'w')
        file.write(scoreboard)
        file.close()

if __name__ == '__main__':
    main()